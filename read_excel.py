import subprocess
import sys

# Try to use openpyxl if available, otherwise install it
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--target=D:/python-pkgs"])
    sys.path.insert(0, "D:/python-pkgs")
    import openpyxl

import os
os.environ["TMPDIR"] = "D:/claude-tmp"
os.environ["TEMP"] = "D:/claude-tmp"
os.environ["TMP"] = "D:/claude-tmp"

# Read the Excel file
xlsx_path = r"C:\Users\Lenovo\Desktop\蓝染纹样生成数据库.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

print("=== Sheet names ===")
for name in wb.sheetnames:
    print(f"  - {name}")

# Read each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")

    # Print headers (first 3 rows)
    for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=False):
        values = [str(cell.value)[:50] if cell.value is not None else "" for cell in row]
        print(f"  | {' | '.join(values)} |")

    # Print first 15 data rows
    print(f"  --- Data rows (up to 15) ---")
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=min(18, ws.max_row), values_only=True), start=4):
        values = [str(v)[:80] if v is not None else "" for v in row]
        print(f"  Row {row_idx}: {' | '.join(values)}")

# Check for images in the workbook
print("\n=== Images ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    images = ws._images
    print(f"  Sheet '{sheet_name}': {len(images)} images")
    for i, img in enumerate(images):
        print(f"    Image {i+1}: anchor={img.anchor}, width={img.width}, height={img.height}")
